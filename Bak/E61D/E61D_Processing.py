import os, random, time
import prophet
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import datetime
import openpyxl
from scipy.stats import mode

def year_week_cal(tmp_set_, col_name_, format_):
    tmp_cal_year = list(pd.to_datetime(tmp_set_[col_name_], format=format_).dt.isocalendar().year.values)
    tmp_cal_ww   = pd.to_datetime(tmp_set_[col_name_], format=format_).dt.isocalendar().week.values
    tmp_cal_ww1 = list(map(lambda x : str(tmp_cal_ww[x]) if divmod(tmp_cal_ww[x], 10)[0] > 0 else str(0)+str(tmp_cal_ww[x]), range(0, len(tmp_cal_year))))
    tmp_cal_date = list(map(lambda x : str(tmp_cal_year[x]) + tmp_cal_ww1[x], range(0, len(tmp_cal_year))))
    return tmp_cal_date

f_list = os.listdir('./data/')
f_list.sort()
production_file = [i for i, v in enumerate(f_list) if 'PRODUCTION' in v.split('_')]
sale_file = [i for i, v in enumerate(f_list) if 'MKT' in v.split('_')]
inven_file = [i for i, v in enumerate(f_list) if 'INVENTORY' in v.split('_')]
#print(production_file)#, sale_file, inven_file)
#print(f_list[production_file[0]])

for i in range(0, len(production_file)):
    tmp_file_raw = openpyxl.load_workbook('./data/'+f_list[production_file[i]])
    if 'RESULT' in f_list[production_file[i]].split('_'):    
        sheet_name = '02.Detail_생산계획 준수율'
        tmp_file_pd = pd.DataFrame(tmp_file_raw[sheet_name].values).copy()
        tmp_col_name1 = list(tmp_file_pd.iloc[5,0:])
        tmp_file = tmp_file_pd.iloc[6:, :].copy().reset_index(drop=True)
        tmp_file.columns = tmp_col_name1
        tmp_file1 = tmp_file[tmp_file['Material'].str.slice(0,12)=='ACEN1060I-B1'].copy().reset_index(drop=True)
        tmp_file1 = tmp_file1[['Plan Date', 'Plan Week', 'Result Qty']] #'Plan Date', 
        tmp_file1['WW'] = pd.to_datetime(tmp_file1['Plan Date']).dt.isocalendar().week
        tmp_file1 = tmp_file1[['Plan Week', 'WW', 'Result Qty']] #'Plan Date', 
        tmp_group = tmp_file1.groupby(['Plan Week', 'WW'])
        tmp_g_sum = pd.DataFrame(tmp_group.sum())
        tmp_g_sum.reset_index(inplace=True)
        tmp_g_cnt = pd.DataFrame(tmp_group.count())
        tmp_g_cnt.reset_index(inplace=True)
        tmp_g = pd.merge(tmp_g_sum, tmp_g_cnt, on=['Plan Week', 'WW'])
    else:    
        sheet_name = 'DynamicResult'
        tmp_file_pd = pd.DataFrame(tmp_file_raw[sheet_name].values).copy()
        tmp_col_name1 = list(tmp_file_pd.iloc[0,0:])
        tmp_file = tmp_file_pd.iloc[1:, :].copy().reset_index(drop=True)
        tmp_file.columns = tmp_col_name1    
        tmp_file1 = tmp_file[['PLAN_START_YYYYMMDD', 'YYYYMMDD', 'QTY']].copy() #'Plan Date', 
        tmp_plan_date_= year_week_cal(tmp_file1, 'PLAN_START_YYYYMMDD', '%Y%m%d')
        tmp_date_= year_week_cal(tmp_file1, 'YYYYMMDD', '%Y%m%d')
        tmp_file1['Plan Date'] = tmp_plan_date_
        tmp_file1['Date']=tmp_date_
        tmp_file1 = tmp_file1[['Plan Date', 'Date', 'QTY']] #'PLAN_START_YYYYMMDD', 'YYYYMMDD', 
        tmp_group = tmp_file1.groupby(['Plan Date', 'Date'])
        tmp_g_sum = pd.DataFrame(tmp_group.sum())
        tmp_g_sum.reset_index(inplace=True)
        tmp_g_cnt = pd.DataFrame(tmp_group.count())
        tmp_g_cnt.reset_index(inplace=True)
        tmp_g = pd.merge(tmp_g_sum, tmp_g_cnt, on=['Plan Date', 'Date'])
    print('Save Complete : ', sheet_name+'.csv')
    tmp_g.to_csv(sheet_name+".csv", index=False)

for i in range(0, len(inven_file)):
    tmp_file_raw = openpyxl.load_workbook('./data/'+f_list[inven_file[i]])
    sheet_name = '02. 일별 PSI 트렌드' #tmp_file_raw.sheetnames
    tmp_file_pd = pd.DataFrame(tmp_file_raw[sheet_name].values).copy()
    tmp_col_name1 = list(tmp_file_pd.iloc[0,0:])
    tmp_file = tmp_file_pd.iloc[1:,:].copy().reset_index(drop=True)
    tmp_file.columns = tmp_col_name1
    tmp_day = pd.to_datetime(tmp_file['연월일'], format='%Y년%m월%d일').dt.isocalendar().day
    tmp_file1 = tmp_file[(tmp_file['자재'].str.slice(0,12)=='ACEN1060I-B1') & (tmp_day == 7)].copy()
    inven_filter_ = [['3', 'C'], ['3','T'],['5','0']]
    for k in range(0, len(inven_filter_)):
        tmp_store1, tmp_store2 = inven_filter_[k][0], inven_filter_[k][1]
        tmp_file1_sub1 = tmp_file1[(tmp_file1['저장위치'].str.slice(0,1) == tmp_store1)&(tmp_file1['저장위치'].str.slice(3,4) == tmp_store2)].copy()
        tmp_week_ = year_week_cal(tmp_file1_sub1, '연월일', '%Y년%m월%d일')
        tmp_file1_sub1['Inven Week'] = tmp_week_
        tmp_file1_sub1 = tmp_file1_sub1[['Inven Week','재고수량']]
        tmp_group = tmp_file1_sub1.groupby(['Inven Week'])
        tmp_sum = pd.DataFrame(tmp_group.sum())
        tmp_sum.reset_index(inplace=True)    
        tmp_sum.to_csv('Inven_%sxx%s.csv'%(tmp_store1, tmp_store2), index=False)
        print('Save Complete : Inven_%sxx%s.csv'%(tmp_store1, tmp_store2))

for f in range(0, len(sale_file)):

    tmp_file_raw = openpyxl.load_workbook('./data/'+f_list[sale_file[f]])
    sheet_name = 'Mkt Fcst Waterfall'
    tmp_file = pd.DataFrame(tmp_file_raw[sheet_name].values)
    ind_col = np.where(tmp_file.iloc[0, :] == 'Measure')[0][0]
    col_name1 = list(tmp_file.iloc[0,:(ind_col+1)])
    tmp_col_name2 = list(tmp_file.iloc[0,(ind_col+1):])
    tmp_col_name2_ = []
    if None in tmp_col_name2:
        for i in range(0, len(tmp_col_name2)):    
            if tmp_col_name2[i] is not None:
                tmp_base = tmp_col_name2[i]
            tmp_col_name2_.append(tmp_base)
    else:
        tmp_col_name2_ = tmp_col_name2

    tmp_col_name3 = list(tmp_file.iloc[1,(ind_col+1):])

    col_name2 = []
    for i in range(0, len(tmp_col_name3)):
        if tmp_col_name3[i] == "SUM":
            col_name2.append(tmp_col_name3[i])
        else:
            col_name2.append(tmp_col_name2_[i][:4]+tmp_col_name3[i][1:])
    col_name = col_name1 + col_name2
    tmp_file1 = tmp_file.iloc[3:,:].reset_index(drop=True).copy()
    tmp_file1.columns = col_name
    if 'SUM' in col_name:
        rm_col1 = np.array([i for i, val in enumerate(col_name) if val == 'SUM'])
        tmp_file1.drop(tmp_file1.iloc[:, rm_col1], axis=1, inplace=True)
    keep_col_name = list(['Version']) + list(tmp_file1.columns[(ind_col+1):])
    tmp_file1 = tmp_file1[keep_col_name].copy()

    tmp_file1_title = tmp_file1.columns.unique()
    tmp_file2 = tmp_file1['Version'].copy()

    for k in range(1, len(tmp_file1_title)):
        tmp_i = [i for i, val in enumerate(tmp_file1.columns) if val == tmp_file1_title[k] ]
        #print(k, tmp_i, tmp_file1_title[k])
        if len(tmp_i) > 1:
            #tmp_file2.append(list(())
            tmp_file2 = pd.concat([tmp_file2, tmp_file1.iloc[:,tmp_i].fillna(0).sum(axis=1)], axis=1)
        else :
            tmp_file2 = pd.concat([tmp_file2, tmp_file1.iloc[:,tmp_i].fillna(0)], axis=1)

    tmp_file2.columns = tmp_file1_title
    tmp_file2['Date'] = year_week_cal(tmp_file2, 'Version', 'MF_W_%Y%m%d_V001')
    tmp_file2 = tmp_file2.sort_values(by=['Date'], axis=0)
    tmp_group = tmp_file2.groupby('Date')
    tmp_g_sum = pd.DataFrame(tmp_group.sum())
    tmp_g_sum.reset_index(inplace=True)
    tmp_g_sum.to_csv(f_list[sale_file[f]][10:].split('.')[0]+'.csv', index=False)
''''''